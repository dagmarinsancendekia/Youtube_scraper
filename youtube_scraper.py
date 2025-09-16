import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import ttk, messagebox

def setup_driver():
    """
    Set up the Chrome WebDriver with necessary options for scraping.
    """
    options = Options()
    # options.add_argument("--headless")  # Uncomment for headless mode
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--ignore-ssl-errors=yes")
    options.add_argument("--ignore-certificate-errors")
    options.add_argument("--disable-web-security")
    options.add_argument("--allow-running-insecure-content")
    options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")
    driver = webdriver.Chrome(options=options)
    return driver

def scrape_youtube(search_query, num_pages=1):
    """
    Scrape YouTube search results for videos based on the search query.

    Args:
        search_query (str): The keyword to search for on YouTube.
        num_pages (int): Number of pages to scrape.

    Returns:
        list: List of dictionaries containing video data.
    """
    driver = setup_driver()
    data = []

    try:
        for page in range(num_pages):
            start = page * 20
            url = f"https://www.youtube.com/results?search_query={search_query}&page={page+1}&sp=EgIQAQ%253D%253D"
            driver.get(url)
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.TAG_NAME, "ytd-item-section-renderer"))
            )
            time.sleep(3)  # Wait for dynamic content

            soup = BeautifulSoup(driver.page_source, 'html.parser')
            videos = soup.find_all('ytd-video-renderer', {'class': 'style-scope ytd-item-section-renderer'})
            print(f"Found {len(videos)} videos on page {page+1}")

            for video in videos:
                title_elem = video.find('a', id='video-title')
                title = title_elem.text.strip() if title_elem else 'N/A'

                channel_elem = video.find('ytd-channel-name').find('a') if video.find('ytd-channel-name') else None
                channel = channel_elem.text.strip() if channel_elem else 'N/A'

                meta_spans = video.find_all('span', class_='style-scope ytd-video-meta-block')
                if len(meta_spans) >= 2:
                    views = meta_spans[0].text.strip()
                    upload_date = meta_spans[1].text.strip()
                else:
                    views = 'N/A'
                    upload_date = 'N/A'

                data.append({
                    'Title': title,
                    'Channel': channel,
                    'Views': views,
                    'Upload Date': upload_date
                })

        driver.quit()
        return data

    except Exception as e:
        print(f"Error: {e}")
        driver.quit()
        return []

def save_to_excel(data, filename='youtube_data.xlsx'):
    """
    Save the scraped data to an Excel file with formatting.

    Args:
        data (list): List of dictionaries containing video data.
        filename (str): Name of the output Excel file.
    """
    df = pd.DataFrame(data)
    df.to_excel(filename, index=False, engine='openpyxl')

    # Format Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "YouTube Data"

    # Write headers
    headers = list(df.columns)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Write data
    for row_num, row in enumerate(df.values, 2):
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal='left')

    # Adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 30

    wb.save(filename)

def create_gui():
    """
    Create a GUI for the YouTube scraper using Tkinter.
    """
    root = tk.Tk()
    root.title("YouTube Scraper")
    root.geometry("800x600")

    # Labels and entries
    tk.Label(root, text="Search Query:").grid(row=0, column=0, padx=10, pady=10)
    query_entry = tk.Entry(root, width=50)
    query_entry.grid(row=0, column=1, padx=10, pady=10)

    tk.Label(root, text="Number of Pages:").grid(row=1, column=0, padx=10, pady=10)
    pages_entry = tk.Entry(root, width=10)
    pages_entry.grid(row=1, column=1, padx=10, pady=10, sticky='w')

    scraped_data = []

    # Treeview for data
    tree = ttk.Treeview(root, columns=('Title', 'Channel', 'Views', 'Date'), show='headings', height=15)
    tree.heading('Title', text='Title')
    tree.heading('Channel', text='Channel')
    tree.heading('Views', text='Views')
    tree.heading('Date', text='Upload Date')
    tree.column('Title', width=200)
    tree.column('Channel', width=150)
    tree.column('Views', width=100)
    tree.column('Date', width=100)
    tree.grid(row=3, column=0, columnspan=2, padx=10, pady=10)

    # Scrollbar for treeview
    scrollbar = ttk.Scrollbar(root, orient=tk.VERTICAL, command=tree.yview)
    tree.configure(yscroll=scrollbar.set)
    scrollbar.grid(row=3, column=2, sticky='ns')

    def scrape():
        nonlocal scraped_data
        query = query_entry.get()
        try:
            pages = int(pages_entry.get())
        except ValueError:
            messagebox.showerror("Error", "Number of pages must be an integer")
            return
        if not query:
            messagebox.showerror("Error", "Search query cannot be empty")
            return
        # Clear previous data
        for item in tree.get_children():
            tree.delete(item)
        # Scrape
        data = scrape_youtube(query, pages)
        scraped_data = data
        if data:
            for row in data:
                tree.insert('', 'end', values=(row['Title'], row['Channel'], row['Views'], row['Upload Date']))
        else:
            messagebox.showinfo("Info", "No data found")

    def save():
        if scraped_data:
            save_to_excel(scraped_data)
            messagebox.showinfo("Success", "Data saved to youtube_data.xlsx")
        else:
            messagebox.showerror("Error", "No data to save")

    # Buttons
    scrape_btn = tk.Button(root, text="Scrape", command=scrape, bg='lightblue')
    scrape_btn.grid(row=2, column=0, padx=10, pady=10)

    save_btn = tk.Button(root, text="Save to Excel", command=save, bg='lightgreen')
    save_btn.grid(row=2, column=1, padx=10, pady=10, sticky='w')

    root.mainloop()

if __name__ == "__main__":
    try:
        create_gui()
    except ImportError:
        # Fallback to CLI if Tkinter not available
        search_query = input("Masukkan kata kunci pencarian YouTube: ")
        num_pages = int(input("Masukkan jumlah halaman: "))
        data = scrape_youtube(search_query, num_pages)
        if data:
            save_to_excel(data)
            print("Data berhasil disimpan ke youtube_data.xlsx")
        else:
            print("Tidak ada data yang ditemukan.")
